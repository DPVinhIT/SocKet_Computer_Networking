import socket
import threading
import os
import tkinter as tk
from tkinter import messagebox, ttk, Frame,Text, Button, VERTICAL, WORD, Label
from tkinter import END
from openpyxl import Workbook, load_workbook
import time
import logging

PORT = 12345
HEADER = 1024
FORMAT = 'utf-8'
DISCONNECT_MESSAGE = "!DISCONNECT"
FILE_UPLOAD_FILE_MESSAGE = "!UPLOAD_FILE"
FILE_UPLOAD_FOLDER_MESSAGE="!UPLOAD_FOLDER"
FILE_LIST_REQUEST = "!LIST"
FILE_DOWNLOAD_REQUEST = "!DOWNLOAD"
SERVER = "0.0.0.0"
ADDR = (SERVER, PORT)

all_connections = []

#File hoạt động của server
log_directory = "server_data/PRIVATE"
os.makedirs(log_directory, exist_ok=True)
log_file_path = os.path.join(log_directory, "server_log.txt")
logging.basicConfig(
    filename=log_file_path,  # Sử dụng đường dẫn đầy đủ
    level=logging.INFO,      # Mức độ log
    format="%(asctime)s - %(levelname)s - %(message)s",  # Định dạng log
    datefmt='[%d/%m/%Y %H:%M:%S]',
    filemode="a"             # Append mode (ghi thêm vào file)
)

# Folder của server
folder_path = "server_data"
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
public_folder = os.path.join(folder_path, "PUBLIC")
private_folder = os.path.join(folder_path, "PRIVATE")
if not os.path.exists(public_folder):
    os.makedirs(public_folder)
if not os.path.exists(private_folder):
    os.makedirs(private_folder)

connected_clients = 0  # Biến đếm số lượng kết nối

# Hàm để server tạo socket kết nối 
def start_server():
    global server
    # Tạo socket với IPv4 và giao thức TCP
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    # Gắn địa chỉ và port cho socket 
    server.bind(ADDR)
    
    insert_chat_box("[STARTING] Server is starting...","green")
    server.listen()
    insert_chat_box(f"[LISTENING] Server is listening on {SERVER}","green")
    logging.info("[Server Start!!]") #Ghi vào file log
    
    show_connecting_button.config(state="normal")
    end_server_button.config(state="normal")
    start_server_button.config(state="disabled")
    
    # Xử lí đa luồng tránh xung đột với giao diện 
    handle_server = threading.Thread(target=server_listen)
    handle_server.daemon = True # Dừng Thread khi đóng giao diện
    handle_server.start()
    
def server_listen():
    while True:
        conn, addr = server.accept()
        all_connections.append(conn)
        # Xử lí đa luồng từng client là 1 luồng khác nhau 
        thread = threading.Thread(target=handle_client, args=(conn, addr))
        thread.daemon =True
        thread.start()
##################################################################################################################################
# Đăng nhập 
def create_or_load_user_db():
    # Đường dẫn đến thư mục PRIVATE
    os.makedirs(private_folder, exist_ok=True)  # Tạo thư mục nếu chưa tồn tại

    # Đường dẫn đầy đủ đến file user_data.xlsx
    file_name = os.path.join(private_folder, "user_data.xlsx")

    try:
        if not os.path.exists(file_name):
            # Tạo file mới và thêm tiêu đề
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Username", "Password"])  # Thêm tiêu đề
            wb.save(file_name)  # Lưu file
        else:
            # Mở file nếu đã tồn tại
            wb = load_workbook(file_name)
        return wb
    except Exception as e:
        print(f"Lỗi khi mở hoặc tạo file: {e}")
        return None
# Đăng ký tài khoản, lưu vào file Excel trong thư mục PRIVATE.
def register(username, password, addr):
    # Đường dẫn tới file trong thư mục PRIVATE
    os.makedirs(private_folder, exist_ok=True)  # Đảm bảo thư mục tồn tại
    file_name = os.path.join(private_folder, "user_data.xlsx")

    # Tải hoặc tạo file Excel
    wb = create_or_load_user_db()
    sheet = wb.active

    # Kiểm tra nếu username đã tồn tại
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            print(f"[Client {addr}] Đăng kí tài khoản không thành công. Tài khoản đã tồn tại.")
            return False

    # Nếu không tồn tại, thêm tài khoản mới vào file
    sheet.append([username, password])
    wb.save(file_name)  # Lưu vào file với đường dẫn đúng
    print(f"[Client {addr}]Đăng ký thành công!")
    return True

# Đăng nhập, kiểm tra tài khoản trong file Excel tại thư mục PRIVATE.
def login(username, password,addr):
    # Đường dẫn tới file trong thư mục PRIVATE
    os.makedirs(private_folder, exist_ok=True)  # Đảm bảo thư mục tồn tại
    file_name = os.path.join(private_folder, "user_data.xlsx")

    # Tải file Excel
    wb = create_or_load_user_db()
    sheet = wb.active

    # Kiểm tra tài khoản và mật khẩu
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            print(f"[Client {addr}] Đăng nhập thành công!")
            return True
    print(f"[Client {addr}] Đăng nhập thất bại!")
##################################################################################################################################
# Xử lí giao tiếp với Client 

# Hàm gửi 1 tin nhắn 
def send_message(conn,msg):
    message = msg.encode(FORMAT)
    msg_length = len(message)
    send_length = str(msg_length).encode(FORMAT)
    send_length += b' ' * (HEADER - len(send_length))
    # Gửi chiều dài tin nhắn là HEADER trước
    try :
        conn.send(send_length)
            # Rồi mới gửi thông điệp
        conn.send(message)
    except Exception as e:
        print("Lỗi khi gửi tin nhắn: {e}")

# Hàm nhận 1 tin nhắn  
def receive_message(conn):
    # Nhận tin nhắn chiều dài là HEADER rồi decode 
    msg_length = conn.recv(HEADER)
    if not msg_length:
        return None
    try:
        msg_length = msg_length.decode(FORMAT)
    except Exception as e:
        print(f"Lỗi chiều dài tin nhắn: {e}")
    msg_length = int(msg_length)  
    # Sau đó mới nhận tin nhắn thực sự 
    msg = conn.recv(msg_length)
    try:
        msg = msg.decode(FORMAT)
    except Exception as e:
        print(f"Lỗi nhận tin nhắn: {e}")
    return msg

# Hàm để gửi data dữ liệu 
def send_data(conn,data):
    try:
        conn.send(data)
        return True
    except Exception as e:
        print(f"Lỗi khi gửi dữ liệu: {e}")
        return False

##################################################################################################################################
# Xử lí file và foler của Client yêu cầu 

######################
# DOWNLOAD FILE 
# Xây dựng cây thư mục với nhiều nhánh từ thư mục gốc, bao gồm cả các tệp và thư mục con.
def build_folder_tree(folder_path):
    base_folder = os.path.basename(folder_path)
    folder_tree = {base_folder: {}}

    # Duyệt qua thư mục gốc và các thư mục con
    for root, dirs, files in os.walk(folder_path):
        # Tính toán đường dẫn tương đối từ thư mục gốc
        relative_path = os.path.relpath(root, folder_path)
        
        # Nếu đường dẫn là thư mục gốc thì tạo node cho thư mục gốc
        parts = relative_path.split(os.sep) if relative_path != '.' else []

        # Duyệt qua các phần tử của đường dẫn tương đối và xây dựng cây
        node = folder_tree[base_folder]  # Bắt đầu từ thư mục cuối cùng
        for part in parts:
            node = node.setdefault(part, {})

        # Thêm các tệp vào thư mục tương ứng
        for file in files:
            node[file] = None  # Các tệp sẽ không có giá trị con, nên ta gán giá trị None

        # Thêm thư mục con vào nếu có
        for dir in dirs:
            if dir not in node:
                node[dir] = {}

    return folder_tree 

# Hàm gửi dữ liệu của file 
def handle_download_file(conn,addr,msg):
    filename = msg.split(" ", 1)[1]
    # Xác định đường dẫn tuyệt đối của file trong thư mục server_datadata
    file_path = os.path.abspath(os.path.join(folder_path, filename))
    # Đường dẫn tuyệt đối của thư mục PUBLIC
    public_folder_path = os.path.abspath(os.path.join(folder_path, "PUBLIC"))
    # Đường dẫn tuyệt đối của thư mục PRIVATE
    private_folder_path = os.path.abspath(os.path.join(folder_path, "PRIVATE"))
    # So sánh các đường dẫn của file với thư mục PUBLIC
    logging.info(f"[Download file] \"{filename}\" from client {addr}")
    if not os.path.relpath(file_path, public_folder_path).startswith(os.pardir):
    # Nếu không nằm ngoài thư mục PUBLIC
        if os.path.exists(file_path):
            # Gửi tên file
            send_message(conn,str(filename.split("/")[-1]))
            file_size = os.path.getsize(file_path)
            # Gửi kích thước file 
            send_message(conn,str(file_size))
            
            with open(file_path, "rb") as file:
                while chunk := file.read(1024):  # Đọc tệp theo từng phần 1024 byte
                    if send_data(conn,chunk):
                        continue
                    else:
                        insert_chat_box(f"Lỗi khi gửi file hoặc kết nối tới {conn} đã bị đóng","red")
                        break
                                
            logging.info(f"[Download file] Successful: \"{filename}\" from client {addr}")
        else:
            send_message(conn,"File not found.")
            logging.error(f"[Download file] Unsuccessful: \"{filename}\" from client {addr}")
    else:
    # Nếu file yêu cầu nằm ngoài thư mục PUBLIC (tức là trong PRIVATE hoặc thư mục khác)
        if os.path.commonpath([file_path, private_folder_path]) == private_folder_path:
            send_message(conn,"File not found.")
            logging.warning(f"[Download file] Unsuccessful: \"{filename}\" from client {addr}")

# Hàm xử lí nếu tên file bị trùng 
def get_unique_filename(file_path):
    if not os.path.exists(file_path):
        return file_path

    base, ext = os.path.splitext(file_path)
    counter = 1
    while True:
        new_file_path = f"{base}({counter}){ext}"
        if not os.path.exists(new_file_path):
            return new_file_path
        counter += 1
##############################
# UPLOAD FILE 
# Hàm nhận dữ liệu file từ client 
def handle_file_upload(conn, addr, msg):
    file_name = msg.split(" ", 1)[1]
    file_length = int(conn.recv(HEADER).decode(FORMAT)) 
    
    logging.info(f"[Upload file] \"{file_name}\" from client {addr}")
    
    # Xác định đường dẫn file 
    file_path = os.path.join(public_folder, file_name)
    # Đổi tên file nếu trùng
    unique_file_path = get_unique_filename(file_path)

    # Đảm bảo tất cả thư mục trong đường dẫn tồn tại
    directory = os.path.dirname(unique_file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)  # Tạo tất cả thư mục cha

    # Nhận dữ liệu từ client và lưu vào file
    total_received = 0
    with open(unique_file_path, "wb") as file:
        while total_received < file_length:
            file_data = conn.recv(1024)
            if not file_data:
                break
            total_received += len(file_data)
            file.write(file_data)
    logging.info(f"[Upload file] Successful: \"{unique_file_path}\" from client {addr}")

################################
# UPLOAD FOLER 
# Hàm nhận dữ liệu file trong folder 
def handle_folder_upload(conn, addr):
    try:

        folder_name = receive_message(conn)  # Nhận tên thư mục gốc
        logging.info(f"[UPLOAD FOLDER] \"{folder_name}\" from client {addr}")
        folder_path = os.path.join(public_folder, folder_name)
        
        # Nếu thư mục đã tồn tại, thay đổi tên thư mục bằng cách thêm hậu tố
        if os.path.exists(folder_path):
            base_folder_name = folder_name
            counter = 1
            while os.path.exists(folder_path):
                folder_path = os.path.join(public_folder, f"{base_folder_name} ({counter})")
                counter += 1
        
        # Tạo thư mục trong PUBLIC
        os.makedirs(folder_path, exist_ok=True)
        try: 
            while True:
                # Nhận tên thư mục con (relative path)
                relative_path = receive_message(conn)
                if relative_path == "EOF":  # Nếu gặp tín hiệu kết thúc
                    break
                
                absolute_path = os.path.join(folder_path, relative_path)
                
                # Xác định folder hay filefile
                type = receive_message(conn)

                if type == "FOLDER":
                    if not os.path.exists(absolute_path):
                        os.makedirs(absolute_path)
                elif type == "FILE":
                    
                    file_length = int(conn.recv(HEADER).decode(FORMAT))
                    total_received = 0 
                    with open(absolute_path, "wb") as file:
                        while total_received < file_length:
                            file_data = conn.recv(1024)
                            if not file_data:
                                logging.error(f"Lỗi khi nhận dữ liệu của file: {absolute_path}")
                                break
                            total_received += len(file_data)
                            file.write(file_data)

                    logging.info(f"File \"{absolute_path}\" uploaded successfully.")
        except Exception as e:
            insert_chat_box(f"Lỗi khi nhận dữ liệu của folder: {e}","red")
            logging.error(f"Lỗi trong quá trình upload folder: {e}")
        logging.info(f"[Up load Folder] \"{folder_name}\" uploaded folder successfully from client {addr}")

    except Exception as e:
        logging.error(f"Error during folder upload from {addr}: {e}")
        conn.send("Error during upload.".encode(FORMAT))

# Hàm xử lí 1 client 
def handle_client(conn, addr):
    global connected_clients
    insert_chat_box(f"[NEW CONNECTION] {addr} connected.","green")
    connected_clients += 1
    insert_chat_box(f"[ACTIVE CONNECTIONS] {connected_clients} active connections.","green")
    logging.info(f"[CONNECT!!] from client {addr}")

    logged_in = False  # Trạng thái chưa đăng nhập

    try:
        while True:
                # Nhận tin nhắn yêu cầu từ client 
                msg = receive_message(conn)
                if stop_event.is_set():  # Kiểm tra sự kiện dừng nếu có
                    break
                if msg == DISCONNECT_MESSAGE:
                    break
                # Nếu chưa đăng nhập
                if not logged_in:
                    if msg.startswith("!REGISTER"):
                        try:
                            _, username, password = msg.split(" ", 2)
                            if register(username, password,addr):
                                conn.send("Registration successful.".encode(FORMAT))
                            else:
                                conn.send("Registration failed. Username already exists.".encode(FORMAT))
                        except ValueError:
                            conn.send("Invalid format. Use: !REGISTER <username> <password>".encode(FORMAT))

                    elif msg.startswith("!LOGIN"):
                        try:
                            _, username, password = msg.split(" ", 2)
                            if login(username, password,addr):
                                conn.send("Login successful.".encode(FORMAT))
                                logged_in = True
                            else:
                                conn.send("Login failed.".encode(FORMAT))
                        except ValueError:
                            conn.send("Invalid format. Use: !LOGIN <username> <password>".encode(FORMAT))
                    else:
                        conn.send("Please login or register first.".encode(FORMAT))
                # Xử lý các yêu cầu sau khi đã đăng nhập
                else:
                    #List file
                    if msg.startswith(FILE_LIST_REQUEST):
                        folder_tree = build_folder_tree(public_folder)
                        folder_data = str(folder_tree)
                        conn.sendall(folder_data.encode(FORMAT))
                        conn.send("EOF".encode(FORMAT))
                    # Upload file
                    elif msg.startswith(FILE_UPLOAD_FILE_MESSAGE):
                        handle_file_upload(conn,addr,msg)
                        
                    # Upload folder
                    elif msg.startswith(FILE_UPLOAD_FOLDER_MESSAGE):
                        handle_folder_upload(conn, addr)
                        
                    # Download file
                    elif msg.startswith(FILE_DOWNLOAD_REQUEST):
                        handle_download_file(conn,addr,msg)

                    # Nếu là tin nhắn 
                    else:
                        insert_chat_box(f"[Client {addr} send]: {msg}","black")
    except Exception as e:
        messagebox.showerror("Lỗi",f"Client {addr} ngắt kết nối\nhoặc\nLỗi handle client: {e}")
    finally:
        all_connections.remove(conn)
        conn.close()
        connected_clients -= 1
        logging.info(f"[DISCONNECT!!] from client {addr}")
        insert_chat_box(f"[DISCONNECT] client {addr} disconnect","red")
        insert_chat_box(f"[ACTIVE CONNECTIONS] {connected_clients} active connections.","red")


##################################################################################################################################
# GIAO DIỆN

# Hàm xử lí khi ấn nút clear text 
def clear_text_box():
    chat_box.config(state="normal")
    chat_box.delete(1.0, END)
    chat_box.config(state="disabled")
# Hàm để in 1 chuỗi lên chat box 
def insert_chat_box(str,color):
    chat_box.config(state="normal")
    chat_box.insert(END,current_time() + ": " + str+"\n",color)
    chat_box.config(state="disabled")
# Hàm tính thời gian hiện tại của máy 
def current_time():
    cur_time = time.strftime("[%d/%m/%Y %H:%M:%S]",time.localtime())
    return str(cur_time)
# Hàm hiển thị các client đang kết nối  
def list_all_connecting():
    if not all_connections:
        insert_chat_box("Không có kết nối","red")
    else: 
        insert_chat_box(f"Có {connected_clients} client đang kết nối là:","green")
        for conn in all_connections:
            try:
                addr=conn.getpeername()
                chat_box.config(state="normal")
                chat_box.insert(END,f" -Client {addr} đang kết nối\n","blue")
                chat_box.config(state="normal")
            except Exception as e:
                chat_box.config(state="normal")
                chat_box.insert(END,f" -Lỗi khi lấy thông tin kết nối: {e}\n","blue")
                chat_box.config(state="disabled")
seconds = 0

# Khởi tạo cửa sổ Tkinter
global server_window
server_window = tk.Tk()
server_window.title("Server Interface")
server_window.geometry("925x500+300+200")
server_window.configure(bg="#34495e")

# Hàm xử lí khi ấn nút đóng cửa số 
def closing_window():
    if messagebox.askokcancel("Thoát", "Bạn có muốn ngắt kết nối ??"):
        try: 
            if server.fileno() != -1: 
                server.close()
        finally:
            server_window.quit()
            server_window.destroy()
            os._exit(0)
server_window.protocol("WM_DELETE_WINDOW", closing_window)

# Hàm xử lí khi ấn nút dừng server 
stop_event = threading.Event()
def end_server():
    try:
        logging.info("[SERVER IS SHUTTING DOWN!!!]")
        stop_event.set()  # Kích hoạt sự kiện dừng

        # Đóng tất cả kết nối client
        for conn in all_connections:
            try:
                conn.close()
            except OSError as e:
                logging.warning(f"Lỗi khi đóng kết nối client: {e}")
        all_connections.clear()

        # Kiểm tra và đóng server
        if server.fileno() != -1:
            try:
                server.close()
                insert_chat_box("[STOPPED] Server đã dừng.", "red")
            except OSError as e:
                logging.error(f"Lỗi khi đóng server: {e}")
                insert_chat_box(f"[ERROR] Lỗi khi đóng server: {e}", "red")
        else:
            insert_chat_box("[ERROR] Server không hợp lệ hoặc đã được đóng trước đó.", "red")
    except Exception as e:
        logging.error(f"Lỗi khi dừng server: {e}")
        messagebox.showerror("Lỗi", f"Lỗi khi dừng server: {e}")
    finally:
        show_connecting_button.config(state="disabled")
        end_server_button.config(state="disabled")
        start_server_button.config(state="normal")

# Xử lí giao diện
# Frame hiện thời gian:
time_frame = Frame(server_window, bg="#34495e")
time_frame.place(relx=0.05, rely=0.05, relwidth=0.9, relheight=0.05)
# Frame thông tin
info_frame = Frame(server_window, bg="#2c3e50", bd=3, relief="solid")
info_frame.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.6)
# Frame Nút
button_frame = Frame(server_window, bg="#34495e")
button_frame.place(relx=0.05, rely=0.72, relwidth=0.9, relheight=0.2)

# Hàm tính thời gian
def current_datetime():
    show_current_time.config(text=f"{current_time()}")
    server_window.after(1000, current_datetime)
def count_time():
    global seconds 
    seconds += 1
    show_runtime.config(text=f"Runtime: {seconds // 3600:02}:{(seconds % 3600) // 60:02}:{seconds % 60:02}")
    server_window.after(1000, count_time)
def show_quantity_connected():
    show_connected_clients.config(text=f"Số lượng đang kết nối: {connected_clients}")
    server_window.after(1000, show_quantity_connected)
# Label thời gian hiển thị
show_current_time = Label(time_frame, text="", fg="white",bg="#34495e",anchor="w", font=("Arial", 14, "bold"))
show_current_time.place(relx=0, rely=0, relheight=1, relwidth=0.25)
show_runtime = Label(time_frame, text="", fg="white",bg="#34495e",anchor="w", font=("Arial", 14, "bold"))
show_runtime.place(relx=0.25, rely=0, relheight=1, relwidth=0.3)
show_connected_clients = Label(time_frame,text="", fg="white",bg="#34495e",anchor="w", font=("Arial", 14, "bold"))
show_connected_clients.place(relx=0.5, rely=0, relheight=1, relwidth=0.3)

current_datetime()
count_time()
show_quantity_connected()

# Chat box để hiển thị thông báo và tin nhắn 
chat_box = Text(info_frame, bg="#ecf0f1", fg="#2c3e50", font=("Arial", 14), wrap=WORD,state="disabled")
chat_box.place(relx=0, relheight=1, relwidth=1)
chat_box.tag_configure("red", foreground="red")
chat_box.tag_configure("green", foreground="green")
chat_box.tag_configure("blue", foreground="blue")
chat_box.tag_configure("black",foreground="black")

# Thanh cuộn của chat box 
scrollbar = ttk.Scrollbar(info_frame, orient=VERTICAL, command=chat_box.yview)
scrollbar.place(relx=0.98, rely=0.004, relheight=0.99, relwidth=0.02)
chat_box.config(yscrollcommand=scrollbar.set)

# Nút start server 
start_server_button = Button(button_frame, text="Start Server",command=start_server
                             , bg="#27ae60", fg="white", font=("Helvetica", 16, "bold"))
start_server_button.place(relx=0, rely=0.3, relwidth=0.22, relheight=0.6)
# Nút end server 
end_server_button = Button(button_frame, text="End Server",command=end_server
                           , bg="#c0392b", fg="white", font=("Helvetica", 16, "bold"),state="disabled")
end_server_button.place(relx=0.25, rely=0.3, relwidth=0.22, relheight=0.6)
# Nút hiển thị các client đang kết nối 
show_connecting_button = Button(button_frame, text="Show All Connecting", command=list_all_connecting
                                , bg="#2980b9", fg="white", font=("Helvetica", 16, "bold"),state="disabled")
show_connecting_button.place(relx=0.51, rely=0.3, relwidth=0.28, relheight=0.6)
# Nút xóa các dữ liệu trong text box 
clear_text_button = Button(button_frame, text="Clear text", command=clear_text_box
                           , bg="#f1c40f", fg="white", font=("Helvetica", 16, "bold"))
clear_text_button.place(relx=0.82, rely=0.3, relwidth=0.18, relheight=0.6)

server_window.mainloop() 
